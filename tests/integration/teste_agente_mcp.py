"""
Teste de integração: agente Gemini + ferramentas MCP para emissão de NFS-e.

Lê a primeira linha com dados da planilha, constrói um prompt em linguagem
natural e executa o loop do agente Gemini chamando as mesmas funções que o
servidor MCP exporia via stdio.

Pré-requisitos:
  - Servidor REST rodando (python -m conta_tools_nfse serve --conf api.conf)
  - GEMINI_API_KEY no ambiente
  - pip install litellm openpyxl

Uso:
  python tests/integration/teste_agente_mcp.py --mcp-conf mcp.conf
  python tests/integration/teste_agente_mcp.py --mcp-conf mcp.conf --dry-run
  python tests/integration/teste_agente_mcp.py --mcp-conf mcp.conf \\
      --planilha examples/edm_teste_campinas.xlsx --linha 2
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

_MODELO = "gemini/gemini-2.5-flash"

# ------------------------------------------------------------------ #
# Schema das ferramentas (espelha o servidor MCP)                     #
# ------------------------------------------------------------------ #

_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "listar_prestadores",
            "description": (
                "Lista as empresas emissoras de NFS-e disponíveis no sistema. "
                "DEVE ser a primeira ferramenta chamada. "
                "Apresente a lista ao usuário e aguarde seleção explícita pelo campo 'id'. "
                "Nunca assuma qual empresa usar sem confirmação do usuário."
            ),
            "parameters": {"type": "object", "properties": {}, "required": []},
        },
    },
    {
        "type": "function",
        "function": {
            "name": "montar_emissao",
            "description": (
                "Valida os dados da NFS-e e monta um resumo para conferência. "
                "NÃO emite a nota. Apresente o resumo ao usuário e aguarde confirmação "
                "explícita antes de chamar confirmar_emissao. "
                "REGRAS: prestador_id deve ser exatamente o 'id' de listar_prestadores; "
                "tomador_cnpj ou tomador_cpf é obrigatório — razão social sozinha não basta; "
                "competencia no formato AAAA-MM; valor_servico com ponto decimal."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "prestador_id":        {"type": "string", "description": "ID exato de listar_prestadores. Nunca inferir."},
                    "tomador_cnpj":        {"type": "string", "description": "CNPJ 14 dígitos (com ou sem pontuação). Obrigatório se tomador_cpf ausente."},
                    "tomador_razao_social":{"type": "string"},
                    "competencia":         {"type": "string", "description": "Formato AAAA-MM, ex: 2026-06"},
                    "discriminacao":       {"type": "string"},
                    "codigo_servico":      {"type": "string"},
                    "valor_servico":       {"type": "string", "description": "Valor em reais com ponto decimal, ex: '450.00'"},
                    "tomador_cpf":         {"type": "string", "default": ""},
                    "codigo_cnae":         {"type": "string", "default": ""},
                    "iss_retido":          {"type": "boolean", "default": False},
                    "valor_iss":           {"type": "string", "default": ""},
                    "tomador_email":       {"type": "string", "default": ""},
                    "tomador_logradouro":  {"type": "string", "default": ""},
                    "tomador_numero":      {"type": "string", "default": ""},
                    "tomador_complemento": {"type": "string", "default": ""},
                    "tomador_bairro":      {"type": "string", "default": ""},
                    "tomador_cep":         {"type": "string", "default": ""},
                    "tomador_uf":          {"type": "string", "default": ""},
                },
                "required": [
                    "prestador_id", "tomador_cnpj", "tomador_razao_social",
                    "competencia", "discriminacao", "codigo_servico", "valor_servico",
                ],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "confirmar_emissao",
            "description": (
                "Emite a NFS-e. Só chame após o usuário confirmar EXPLICITAMENTE o resumo "
                "de montar_emissao. Use os mesmos dados do resumo sem alteração, "
                "incluindo o numero_rps retornado. "
                "Se o retorno incluir aviso_rps, informe o usuário que o RPS foi ajustado."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "prestador_id":        {"type": "string"},
                    "numero_rps":          {"type": "string", "description": "RPS retornado por montar_emissao. Não alterar."},
                    "tomador_cnpj":        {"type": "string"},
                    "tomador_razao_social":{"type": "string"},
                    "competencia":         {"type": "string"},
                    "discriminacao":       {"type": "string"},
                    "codigo_servico":      {"type": "string"},
                    "valor_servico":       {"type": "string"},
                    "tomador_cpf":         {"type": "string", "default": ""},
                    "codigo_cnae":         {"type": "string", "default": ""},
                    "iss_retido":          {"type": "boolean", "default": False},
                    "valor_iss":           {"type": "string", "default": ""},
                    "tomador_email":       {"type": "string", "default": ""},
                    "tomador_logradouro":  {"type": "string", "default": ""},
                    "tomador_numero":      {"type": "string", "default": ""},
                    "tomador_complemento": {"type": "string", "default": ""},
                    "tomador_bairro":      {"type": "string", "default": ""},
                    "tomador_cep":         {"type": "string", "default": ""},
                    "tomador_uf":          {"type": "string", "default": ""},
                },
                "required": [
                    "prestador_id", "numero_rps", "tomador_cnpj", "tomador_razao_social",
                    "competencia", "discriminacao", "codigo_servico", "valor_servico",
                ],
            },
        },
    },
]

_SYSTEM_PROMPT = (
    "Você é um assistente para emissão de NFS-e (Nota Fiscal de Serviços Eletrônica). "
    "REGRAS INVIOLÁVEIS:\n"
    "1. Sempre chame listar_prestadores primeiro. Nunca assuma a empresa emissora.\n"
    "2. Nunca assuma dados do tomador. CNPJ ou CPF é obrigatório — nome sozinho não identifica.\n"
    "3. Sempre chame montar_emissao e apresente o resumo antes de emitir.\n"
    "4. Só chame confirmar_emissao após confirmação explícita do usuário.\n"
    "5. Nunca altere valores, RPS ou dados do tomador sem nova confirmação.\n"
    "Responda em português brasileiro."
)


# ------------------------------------------------------------------ #
# Leitura da planilha                                                  #
# ------------------------------------------------------------------ #


def ler_linha_planilha(caminho: Path, linha: int = 2) -> dict:
    """Lê uma linha da planilha e retorna um dict com os campos normalizados."""
    wb = openpyxl.load_workbook(caminho)
    ws = wb.active

    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    def get(nome: str):
        for h, idx in headers.items():
            if nome.lower() in (h or "").lower():
                return ws.cell(linha, idx).value
        return None

    competencia_raw = get("compet")
    if isinstance(competencia_raw, datetime):
        competencia = competencia_raw.strftime("%Y-%m")
    else:
        competencia = str(competencia_raw or "")
        if "/" in competencia:
            m, a = competencia.split("/")
            competencia = f"{a}-{m.zfill(2)}"

    cnpj_raw = str(get("cnpj do tomador") or "").strip()

    return {
        "numero_rps":          str(get("rps") or ""),
        "competencia":         competencia,
        "discriminacao":       str(get("discrimina") or "").strip(),
        "codigo_servico":      str(get("código do servi") or "").strip(),
        "codigo_cnae":         str(int(get("cnae") or 0) if get("cnae") else ""),
        "valor_servico":       str(get("valor dos servi") or ""),
        "iss_retido":          str(get("iss retido") or "N").upper() == "S",
        "tomador_razao_social":str(get("razão social") or "").strip(),
        "tomador_cnpj":        cnpj_raw,
        "tomador_email":       str(get("e-mail") or "").strip(),
        "tomador_logradouro":  str(get("logradouro") or "").strip(),
        "tomador_numero":      str(get("número do end") or "").strip(),
        "tomador_complemento": str(get("complemento") or "").strip(),
        "tomador_bairro":      str(get("bairro") or "").strip(),
        "tomador_cep":         str(get("cep") or "").strip(),
        "tomador_uf":          str(get("uf") or "").strip(),
    }


def _prompt_a_partir_de_linha(dados: dict) -> str:
    """Constrói o prompt em linguagem natural a partir dos dados da planilha."""
    iss = "retido" if dados["iss_retido"] else "não retido"
    return (
        f"Preciso emitir uma NFS-e com os seguintes dados:\n\n"
        f"Tomador: {dados['tomador_razao_social']}\n"
        f"CNPJ: {dados['tomador_cnpj']}\n"
        f"E-mail: {dados['tomador_email']}\n"
        f"Endereço: {dados['tomador_logradouro']}, {dados['tomador_numero']}"
        + (f", {dados['tomador_complemento']}" if dados['tomador_complemento'] else "")
        + f" — {dados['tomador_bairro']} — CEP {dados['tomador_cep']} — {dados['tomador_uf']}\n\n"
        f"Serviço:\n"
        f"  Código: {dados['codigo_servico']}\n"
        + (f"  CNAE: {dados['codigo_cnae']}\n" if dados['codigo_cnae'] else "")
        + f"  Valor: R$ {dados['valor_servico']}\n"
        f"  ISS: {iss}\n"
        f"  Competência: {dados['competencia']}\n\n"
        f"Discriminação:\n{dados['discriminacao']}\n\n"
        f"Por favor, verifique as empresas emissoras disponíveis, monte o resumo da nota "
        f"e me apresente para confirmação antes de emitir."
    )


# ------------------------------------------------------------------ #
# Loop do agente                                                       #
# ------------------------------------------------------------------ #


def _despachar(nome: str, kwargs: dict, dry_run: bool) -> str:
    """Chama a função MCP correspondente e retorna resultado como string JSON."""
    from conta_tools_nfse.mcp.server import (
        confirmar_emissao,
        listar_prestadores,
        montar_emissao,
    )

    if nome == "listar_prestadores":
        return listar_prestadores()

    if nome == "montar_emissao":
        return montar_emissao(**kwargs)

    if nome == "confirmar_emissao":
        if dry_run:
            return json.dumps({
                "status": "dry-run",
                "numero_nfse": "DRY-RUN",
                "codigo_verificacao": "DRY-RUN",
                "link_consulta": "",
                "arquivo": "",
                "_aviso": "dry-run ativo — nota NÃO foi emitida.",
            }, ensure_ascii=False)
        return confirmar_emissao(**kwargs)

    return json.dumps({"erro": f"Ferramenta desconhecida: {nome}"})


def rodar_agente(prompt: str, dry_run: bool, verbose: bool) -> None:
    import litellm

    litellm.suppress_debug_info = True

    messages: list[dict] = [
        {"role": "system", "content": _SYSTEM_PROMPT},
        {"role": "user", "content": prompt},
    ]

    print("=" * 60)
    print("PROMPT INICIAL:")
    print(prompt)
    print("=" * 60)

    turno = 0
    while True:
        turno += 1
        print(f"\n[turno {turno}] Chamando {_MODELO}...")

        resp = litellm.completion(
            model=_MODELO,
            messages=messages,
            tools=_TOOLS,
            tool_choice="auto",
        )

        msg = resp.choices[0].message
        messages.append(msg.model_dump(exclude_none=True))

        tool_calls = getattr(msg, "tool_calls", None) or []

        if not tool_calls:
            texto = msg.content or "(sem resposta de texto)"
            print("\n[agente]", texto)

            # Se o agente fez uma pergunta, permite que o usuário responda
            if any(p in texto for p in ("?", "confirme", "confirma", "informe", "qual", "por favor")):
                try:
                    resposta = input("\n[você] ").strip()
                    if resposta:
                        messages.append({"role": "user", "content": resposta})
                        continue
                except (EOFError, KeyboardInterrupt):
                    pass
            break

        for tc in tool_calls:
            nome = tc.function.name
            try:
                kwargs = json.loads(tc.function.arguments or "{}")
            except json.JSONDecodeError:
                kwargs = {}

            print(f"\n  → {nome}({', '.join(f'{k}={v!r}' for k, v in kwargs.items() if v)})")

            resultado = _despachar(nome, kwargs, dry_run)

            if verbose:
                try:
                    parsed = json.loads(resultado)
                    print(f"  ← {json.dumps(parsed, ensure_ascii=False, indent=4)}")
                except Exception:
                    print(f"  ← {resultado}")
            else:
                # Mostra apenas resumo do resultado
                try:
                    parsed = json.loads(resultado)
                    if isinstance(parsed, list):
                        print(f"  ← [{len(parsed)} item(s)]")
                    elif isinstance(parsed, dict):
                        chaves = list(parsed.keys())[:4]
                        print(f"  ← {{{', '.join(chaves)}...}}")
                    else:
                        print(f"  ← {resultado[:120]}")
                except Exception:
                    print(f"  ← {resultado[:120]}")

            messages.append({
                "role": "tool",
                "tool_call_id": tc.id,
                "content": resultado,
            })

    print("\n" + "=" * 60)
    print("FIM DO LOOP DO AGENTE")
    if dry_run:
        print("(dry-run ativo — nenhuma nota foi emitida)")
    print("=" * 60)


# ------------------------------------------------------------------ #
# Entry point                                                          #
# ------------------------------------------------------------------ #


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Teste de integração: agente Gemini + ferramentas MCP para emissão de NFS-e.",
    )
    parser.add_argument(
        "--mcp-conf",
        type=Path,
        default=Path("mcp.conf"),
        metavar="CONF",
        help="Arquivo mcp.conf com url e bearer_token da API (default: mcp.conf)",
    )
    parser.add_argument(
        "--planilha",
        type=Path,
        default=Path("examples/edm_teste_campinas.xlsx"),
        metavar="XLSX",
        help="Planilha com dados de teste (default: examples/edm_teste_campinas.xlsx)",
    )
    parser.add_argument(
        "--linha",
        type=int,
        default=2,
        metavar="N",
        help="Linha da planilha a usar como dados (default: 2, primeira linha de dados)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Não emite de verdade — substitui confirmar_emissao por mock",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Exibe JSON completo dos retornos das ferramentas",
    )
    parser.add_argument(
        "--api-key",
        metavar="CHAVE",
        help="Gemini API key (alternativa: variável GEMINI_API_KEY)",
    )
    args = parser.parse_args()

    # Gemini API key
    if args.api_key:
        os.environ["GEMINI_API_KEY"] = args.api_key
    elif not os.environ.get("GEMINI_API_KEY"):
        print("ERRO: GEMINI_API_KEY não definida. Use --api-key ou exporte a variável.", file=sys.stderr)
        return 1

    # Imports opcionais
    try:
        import litellm  # noqa: F401
    except ImportError:
        print("ERRO: litellm não instalado. Execute: pip install litellm", file=sys.stderr)
        return 1

    # Carrega mcp.conf e inicializa as funções MCP
    try:
        from conta_tools_nfse.mcp.conf import carregar_mcp_conf
        from conta_tools_nfse.mcp.server import inicializar

        mcp_conf = carregar_mcp_conf(args.mcp_conf)
        inicializar(mcp_conf.api_url, mcp_conf.bearer_token)
        print(f"API: {mcp_conf.api_url}")
    except Exception as e:
        print(f"ERRO ao carregar {args.mcp_conf}: {e}", file=sys.stderr)
        return 1

    # Lê planilha
    try:
        dados = ler_linha_planilha(args.planilha, args.linha)
        print(f"Planilha: {args.planilha} (linha {args.linha})")
        print(f"Tomador: {dados['tomador_razao_social']} — CNPJ: {dados['tomador_cnpj']}")
    except Exception as e:
        print(f"ERRO ao ler planilha: {e}", file=sys.stderr)
        return 1

    # Monta prompt e executa agente
    prompt = _prompt_a_partir_de_linha(dados)
    rodar_agente(prompt, dry_run=args.dry_run, verbose=args.verbose)
    return 0


if __name__ == "__main__":
    # Garante que o pacote é encontrado ao rodar direto da raiz do repo
    repo_src = Path(__file__).parent.parent.parent / "src"
    if str(repo_src) not in sys.path:
        sys.path.insert(0, str(repo_src))

    sys.exit(main())
