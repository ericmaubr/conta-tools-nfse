"""Testes unitários para leitura da planilha de NFS-e."""

from __future__ import annotations

import tempfile
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from conta_tools_nfse.excel.columns import TODAS_COLUNAS
from conta_tools_nfse.excel.reader import _parse_competencia, _parse_data, ler_planilha_campinas
from conta_tools_nfse.excel.template import criar_template_campinas


# ------------------------------------------------------------------ #
# Helpers de fixture                                                   #
# ------------------------------------------------------------------ #

def _wb_minimo(rows: list[dict]) -> Workbook:
    """Cria um Workbook com as colunas mínimas e as linhas fornecidas."""
    wb = Workbook()
    ws = wb.active
    headers = TODAS_COLUNAS
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(1, col_idx, h)
    for row_idx, dados in enumerate(rows, start=2):
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, dados.get(h, ""))
    return wb


def _linha_valida(**kwargs) -> dict:
    base = {
        "numero_rps": "1",
        "competencia": "01/2026",
        "discriminacao": "Consultoria em TI",
        "codigo_servico": "1.07",
        "valor_servico": 1000.00,
        "tomador_razao_social": "Empresa Teste",
        "tomador_cnpj": "12.345.678/0001-90",
    }
    base.update(kwargs)
    return base


# ------------------------------------------------------------------ #
# Parsers                                                              #
# ------------------------------------------------------------------ #

@pytest.mark.parametrize("entrada,esperado", [
    ("01/2026", "2026-01"),
    ("1/2026", "2026-01"),
    ("12/2025", "2025-12"),
    ("13/2025", ""),          # mês inválido
    ("2026-01", ""),          # formato errado
    ("", ""),
])
def test_parse_competencia(entrada, esperado):
    assert _parse_competencia(entrada) == esperado


@pytest.mark.parametrize("entrada,esperado", [
    ("15/01/2026", "2026-01-15"),
    ("2026-01-15", "2026-01-15"),
    ("", ""),
    ("invalido", ""),
])
def test_parse_data(entrada, esperado):
    assert _parse_data(entrada) == esperado


# ------------------------------------------------------------------ #
# Template                                                             #
# ------------------------------------------------------------------ #

def test_criar_template_gera_arquivo():
    with tempfile.TemporaryDirectory() as tmpdir:
        saida = Path(tmpdir) / "template.xlsx"
        criar_template_campinas(saida)
        assert saida.exists()
        assert saida.stat().st_size > 0


# ------------------------------------------------------------------ #
# Leitura de planilha                                                  #
# ------------------------------------------------------------------ #

def _ler(rows: list[dict]):
    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / "test.xlsx"
        wb = _wb_minimo(rows)
        wb.save(path)
        return ler_planilha_campinas(
            path,
            prestador_cnpj="98765432000111",
            inscricao_municipal="99999",
            cert_path=Path("/fake/cert.pfx"),
            cert_senha="senha",
        )


def test_linha_valida_gera_pedido():
    pedidos, erros = _ler([_linha_valida()])
    assert len(pedidos) == 1
    assert not erros
    assert pedidos[0].discriminacao == "Consultoria em TI"
    assert pedidos[0].valor_servico == Decimal("1000.0")
    assert pedidos[0].municipio_prestacao == "campinas"


def test_competencia_invalida_gera_erro():
    _, erros = _ler([_linha_valida(competencia="99/2026")])
    assert any("competencia" in e for e in erros)


def test_sem_cnpj_nem_cpf_gera_erro():
    _, erros = _ler([_linha_valida(tomador_cnpj="", tomador_cpf="")])
    assert any("tomador_cnpj" in e or "tomador_cpf" in e for e in erros)


def test_campo_obrigatorio_vazio_gera_erro():
    _, erros = _ler([_linha_valida(discriminacao="")])
    assert any("discriminacao" in e for e in erros)


def test_linha_vazia_ignorada():
    pedidos, erros = _ler([{}])
    assert len(pedidos) == 0
    assert not erros


def test_multiplas_linhas():
    pedidos, erros = _ler([
        _linha_valida(numero_rps="1"),
        _linha_valida(numero_rps="2", discriminacao=""),
        _linha_valida(numero_rps="3"),
    ])
    assert len(pedidos) == 2
    assert len(erros) == 1
