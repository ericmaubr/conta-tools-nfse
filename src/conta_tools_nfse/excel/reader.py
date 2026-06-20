"""Leitura e validação da planilha de emissão de NFS-e."""

from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from conta_tools_shared.domain.cnpj import only_digits
from conta_tools_shared.domain.nfse import NfseRequest, PrestadorNfse, TomadorNfse

from conta_tools_nfse.excel.columns import (
    COLUNAS_OBRIGATORIAS,
    COLUNAS_SP_OBRIGATORIAS,
    COLUNAS_TOMADOR_ID,
    DESCRICOES,
)


def _limpar_header(h: str) -> str:
    """Remove acentos, asteriscos, parênteses e dicas de formato do cabeçalho."""
    h = re.sub(r"[*()]|mm/aaaa|dd/mm/aaaa|r\$|s/n|s ou n|%", "", h.lower()).strip()
    h = "".join(c for c in unicodedata.normalize("NFD", h) if unicodedata.category(c) != "Mn")
    return h.strip()


# Mapeamento reverso: descrição legível (limpa) → nome programático da coluna
# Permite reconhecer cabeçalhos gerados pelo template (ex: "Número RPS *" → "numero_rps")
_HEADER_TO_COL: dict[str, str] = {
    _limpar_header(desc): col_name for col_name, desc in DESCRICOES.items()
}


_TODAS_COLUNAS_CAMPINAS = [
    "data_emissao", "tomador_inscricao_municipal", "tomador_email",
    "tomador_logradouro", "tomador_numero", "tomador_complemento",
    "tomador_bairro", "tomador_cep", "tomador_municipio_ibge", "tomador_uf",
    "deducoes", "iss_retido", "optante_simples", "natureza_operacao", "codigo_cnae",
]

_TODAS_COLUNAS_SP_OPCIONAIS = [
    "data_emissao", "tomador_inscricao_municipal", "tomador_email",
    "tomador_logradouro", "tomador_numero", "tomador_complemento",
    "tomador_bairro", "tomador_cep", "tomador_municipio_ibge", "tomador_uf",
    "deducoes", "iss_retido", "optante_simples", "tributacao_rps",
]


def _mapear_colunas(ws, colunas_busca: list[str]) -> dict[str, int]:
    """Mapeia nomes de coluna para índice (1-based) usando cabeçalhos da linha 1.

    Reconhece tanto nomes programáticos ("numero_rps") quanto descrições legíveis
    geradas pelo template ("Número RPS *").
    """
    cabecalhos_raw = [
        (ws.cell(1, c).value or "").strip().lower()
        for c in range(1, ws.max_column + 1)
    ]
    cabecalhos = [re.sub(r"[*()]|mm/aaaa|dd/mm/aaaa|r\$|s/n|s ou n|%", "", h).strip() for h in cabecalhos_raw]
    cabecalhos_limpos = [_limpar_header(h) for h in cabecalhos_raw]

    col_map: dict[str, int] = {}
    for idx, (h, h_limpo) in enumerate(zip(cabecalhos, cabecalhos_limpos), start=1):
        for col_name in colunas_busca:
            # 1) Matching por nome programático (backward compat)
            if col_name in h or h in col_name:
                if col_name not in col_map:
                    col_map[col_name] = idx
                break
            # 2) Matching por descrição legível do template (ex: "Número RPS *")
            if _HEADER_TO_COL.get(h_limpo) == col_name:
                if col_name not in col_map:
                    col_map[col_name] = idx
                break
    return col_map


def ler_planilha_campinas(
    caminho: Path,
    prestador_cnpj: str,
    inscricao_municipal: str,
    cert_path: Path,
    cert_senha: str,
) -> tuple[list[NfseRequest], list[str]]:
    """
    Lê a planilha e retorna (pedidos_válidos, lista_de_erros).

    Erros são por linha e não interrompem o processamento das demais.
    """
    wb = load_workbook(caminho, data_only=True)
    ws = wb.active

    col_map = _mapear_colunas(
        ws, COLUNAS_OBRIGATORIAS + COLUNAS_TOMADOR_ID + _TODAS_COLUNAS_CAMPINAS
    )

    faltando = [c for c in COLUNAS_OBRIGATORIAS if c not in col_map]
    if faltando:
        raise ValueError(f"Colunas obrigatórias não encontradas: {', '.join(faltando)}")

    prestador = PrestadorNfse(
        cnpj=only_digits(prestador_cnpj),
        inscricao_municipal=inscricao_municipal,
        cert_path=cert_path,
        cert_senha=cert_senha,
    )

    pedidos: list[NfseRequest] = []
    erros: list[str] = []

    for row_num in range(2, ws.max_row + 1):
        def cel(col_name: str) -> str:
            if col_name not in col_map:
                return ""
            val = ws.cell(row_num, col_map[col_name]).value
            if val is None:
                return ""
            return str(val).strip()

        def cel_decimal(col_name: str, default: str = "0") -> Decimal:
            v = cel(col_name) or default
            v = v.replace(",", ".")
            try:
                return Decimal(v)
            except InvalidOperation:
                raise ValueError(f"Valor inválido em '{col_name}': {v!r}")

        # Pular linhas completamente vazias
        if not any(
            ws.cell(row_num, col_map[c]).value
            for c in COLUNAS_OBRIGATORIAS
            if c in col_map
        ):
            continue

        erros_linha: list[str] = []

        # Validar campos obrigatórios
        for c in COLUNAS_OBRIGATORIAS:
            if not cel(c):
                erros_linha.append(f"campo obrigatório vazio: {c}")

        # Validar identificação do tomador
        cnpj_tom = only_digits(cel("tomador_cnpj"))
        cpf_tom = only_digits(cel("tomador_cpf"))
        if not cnpj_tom and not cpf_tom:
            erros_linha.append("preencha tomador_cnpj ou tomador_cpf")

        # Validar competência
        competencia_raw = cel("competencia")
        competencia_iso = _parse_competencia(competencia_raw)
        if not competencia_iso:
            erros_linha.append(f"competencia inválida: {competencia_raw!r} (esperado MM/AAAA)")

        if erros_linha:
            erros.append(f"Linha {row_num}: {'; '.join(erros_linha)}")
            continue

        # Parsear data de emissão
        data_emissao_raw = cel("data_emissao")
        data_emissao = _parse_data(data_emissao_raw) or date.today().isoformat()

        try:
            valor = cel_decimal("valor_servico")
            deducoes = cel_decimal("deducoes", "0")
        except ValueError as e:
            erros.append(f"Linha {row_num}: {e}")
            continue

        tomador = TomadorNfse(
            razao_social=cel("tomador_razao_social"),
            cnpj=cnpj_tom,
            cpf=cpf_tom,
            inscricao_municipal=cel("tomador_inscricao_municipal"),
            email=cel("tomador_email"),
            logradouro=cel("tomador_logradouro"),
            numero=cel("tomador_numero"),
            complemento=cel("tomador_complemento"),
            bairro=cel("tomador_bairro"),
            cep=only_digits(cel("tomador_cep")),
            municipio_ibge=cel("tomador_municipio_ibge"),
            uf=cel("tomador_uf"),
        )

        iss_retido_raw = cel("iss_retido").upper()
        optante_raw = cel("optante_simples").upper()
        nat_op_raw = cel("natureza_operacao")

        req = NfseRequest(
            id=str(uuid4()),
            prestador=prestador,
            tomador=tomador,
            discriminacao=cel("discriminacao"),
            valor_servico=valor,
            codigo_servico=cel("codigo_servico"),
            municipio_prestacao="campinas",
            competencia=competencia_iso,
            numero_rps=cel("numero_rps"),
            data_emissao=data_emissao,
            deducoes=deducoes,
            iss_retido=iss_retido_raw == "S",
            optante_simples=optante_raw == "S",
            natureza_operacao=int(nat_op_raw) if nat_op_raw.isdigit() else 1,
            codigo_cnae=cel("codigo_cnae"),
        )

        ##print("--- NfseRequest ---")
        ##print(req)
        ##print("-" * 40)

        pedidos.append(req)

    return pedidos, erros


def ler_planilha_sp(
    caminho: Path,
    prestador_cnpj: str,
    inscricao_municipal: str,
    cert_path: Path,
    cert_senha: str,
) -> tuple[list[NfseRequest], list[str]]:
    """
    Lê a planilha SP e retorna (pedidos_válidos, lista_de_erros).

    Diferenças em relação à versão Campinas:
    - Requer coluna `aliquota_servicos` (% do ISS, ex: 5.00 = 5%)
    - Aceita coluna opcional `tributacao_rps` (T/F/J/A/B/M; default "T")
    - Define `serie_rps = "RPS"` conforme padrão SP
    - Não usa `natureza_operacao` nem `codigo_cnae`
    """
    from decimal import Decimal

    wb = load_workbook(caminho, data_only=True)
    ws = wb.active

    obrigatorias_sp = COLUNAS_OBRIGATORIAS + COLUNAS_SP_OBRIGATORIAS
    col_map = _mapear_colunas(
        ws, obrigatorias_sp + COLUNAS_TOMADOR_ID + _TODAS_COLUNAS_SP_OPCIONAIS
    )

    faltando = [c for c in obrigatorias_sp if c not in col_map]
    if not (set(COLUNAS_TOMADOR_ID) & set(col_map)):
        faltando.append("tomador_cnpj ou tomador_cpf")
    if faltando:
        raise ValueError(f"Colunas obrigatórias não encontradas: {', '.join(faltando)}")

    prestador = PrestadorNfse(
        cnpj=only_digits(prestador_cnpj),
        inscricao_municipal=inscricao_municipal,
        cert_path=cert_path,
        cert_senha=cert_senha,
    )

    pedidos: list[NfseRequest] = []
    erros: list[str] = []

    for row_num in range(2, ws.max_row + 1):
        def cel(col_name: str) -> str:
            if col_name not in col_map:
                return ""
            val = ws.cell(row_num, col_map[col_name]).value
            return "" if val is None else str(val).strip()

        def cel_decimal(col_name: str, default: str = "0") -> Decimal:
            v = (cel(col_name) or default).replace(",", ".")
            try:
                return Decimal(v)
            except InvalidOperation:
                raise ValueError(f"Valor inválido em '{col_name}': {v!r}")

        if not any(
            ws.cell(row_num, col_map[c]).value
            for c in COLUNAS_OBRIGATORIAS
            if c in col_map
        ):
            continue

        erros_linha: list[str] = []

        for c in obrigatorias_sp:
            if not cel(c):
                erros_linha.append(f"campo obrigatório vazio: {c}")

        cnpj_tom = only_digits(cel("tomador_cnpj"))
        cpf_tom = only_digits(cel("tomador_cpf"))
        if not cnpj_tom and not cpf_tom:
            erros_linha.append("preencha tomador_cnpj ou tomador_cpf")

        competencia_raw = cel("competencia")
        competencia_iso = _parse_competencia(competencia_raw)
        if not competencia_iso:
            erros_linha.append(f"competencia inválida: {competencia_raw!r} (esperado MM/AAAA)")

        if erros_linha:
            erros.append(f"Linha {row_num}: {'; '.join(erros_linha)}")
            continue

        data_emissao = _parse_data(cel("data_emissao")) or date.today().isoformat()

        try:
            valor = cel_decimal("valor_servico")
            deducoes = cel_decimal("deducoes", "0")
            aliquota = cel_decimal("aliquota_servicos", "0")
        except ValueError as e:
            erros.append(f"Linha {row_num}: {e}")
            continue

        tomador = TomadorNfse(
            razao_social=cel("tomador_razao_social"),
            cnpj=cnpj_tom,
            cpf=cpf_tom,
            inscricao_municipal=cel("tomador_inscricao_municipal"),
            email=cel("tomador_email"),
            logradouro=cel("tomador_logradouro"),
            numero=cel("tomador_numero"),
            complemento=cel("tomador_complemento"),
            bairro=cel("tomador_bairro"),
            cep=only_digits(cel("tomador_cep")),
            municipio_ibge=cel("tomador_municipio_ibge"),
            uf=cel("tomador_uf"),
        )

        trib_raw = cel("tributacao_rps").upper() or "T"
        iss_retido_raw = cel("iss_retido").upper()
        optante_raw = cel("optante_simples").upper()

        req = NfseRequest(
            id=str(uuid4()),
            prestador=prestador,
            tomador=tomador,
            discriminacao=cel("discriminacao"),
            valor_servico=valor,
            codigo_servico=cel("codigo_servico"),
            municipio_prestacao="sao_paulo",
            competencia=competencia_iso,
            numero_rps=cel("numero_rps"),
            serie_rps="RPS",
            data_emissao=data_emissao,
            deducoes=deducoes,
            iss_retido=iss_retido_raw == "S",
            optante_simples=optante_raw == "S",
            tributacao_rps=trib_raw,
            aliquota_servicos=aliquota,
        )
        pedidos.append(req)

    return pedidos, erros


def _parse_competencia(valor: str) -> str:
    """Converte MM/AAAA → AAAA-MM. Aceita também datas ISO (YYYY-MM-DD) do Excel."""
    if not valor:
        return ""
    v = valor.strip()
    # Formato esperado pelo usuário: MM/AAAA ou MM-AAAA
    m = re.fullmatch(r"(\d{1,2})[/\-](\d{4})", v)
    if m:
        mes, ano = int(m.group(1)), int(m.group(2))
        if 1 <= mes <= 12:
            return f"{ano}-{mes:02d}"
    # Formato retornado pelo openpyxl quando a célula é data Excel: YYYY-MM-DD [HH:MM:SS]
    m = re.match(r"(\d{4})-(\d{2})-\d{2}", v)
    if m:
        ano, mes = int(m.group(1)), int(m.group(2))
        if 1 <= mes <= 12:
            return f"{ano}-{mes:02d}"
    return ""


def _parse_data(valor: str) -> str:
    """Converte DD/MM/AAAA → AAAA-MM-DD. Retorna '' se inválido."""
    if not valor:
        return ""
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(valor.strip(), fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def salvar_resultado(
    caminho_original: Path,
    caminho_saida: Path,
    resultados: list[tuple],
) -> None:
    """
    Acrescenta colunas de resultado (status, numero_nfse, etc.) à planilha original.
    resultados: lista de (NfseRequest, NfseResult|None, erro_str|None)
    """
    from openpyxl.styles import Font, PatternFill

    wb = load_workbook(caminho_original)
    ws = wb.active

    # Encontrar próxima coluna vazia após os dados
    next_col = ws.max_column + 1
    headers = ["status", "numero_nfse", "codigo_verificacao", "link_consulta", "erro"]
    cores = {"EMITIDA": "C6EFCE", "ERRO": "FFC7CE"}

    for i, h in enumerate(headers):
        ws.cell(1, next_col + i, h).font = Font(bold=True)

    for row_offset, (req, result, erro) in enumerate(resultados, start=2):
        status = "EMITIDA" if result else "ERRO"
        cor = cores[status]
        fill = PatternFill("solid", fgColor=cor)
        cel_status = ws.cell(row_offset, next_col, status)
        cel_status.fill = fill
        ws.cell(row_offset, next_col + 1, result.numero_nota if result else "")
        ws.cell(row_offset, next_col + 2, result.codigo_verificacao if result else "")
        ws.cell(row_offset, next_col + 3, result.link_consulta if result else "")
        ws.cell(row_offset, next_col + 4, erro or "")

    wb.save(caminho_saida)
