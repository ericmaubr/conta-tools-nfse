"""Geração do template Excel para preenchimento de NFS-e."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from conta_tools_nfse.excel.columns import (
    DESCRICOES,
    EXEMPLO,
    EXEMPLO_SP,
    TODAS_COLUNAS,
    TODAS_COLUNAS_SP,
)

_COR_OBRIGATORIO = "1F4E79"   # azul escuro — colunas obrigatórias
_COR_OPCONAL = "2E75B6"       # azul médio — colunas opcionais
_COR_EXEMPLO = "DEEAF1"       # azul claro — linha de exemplo
_COR_RESULTADO = "595959"     # cinza — colunas de resultado (preenchidas pelo sistema)

_LARGURAS_SP_EXTRAS: dict[str, int] = {
    "aliquota_servicos": 20,
    "tributacao_rps": 24,
}

_LARGURAS: dict[str, int] = {
    "numero_rps": 14,
    "competencia": 18,
    "discriminacao": 50,
    "codigo_servico": 20,
    "valor_servico": 20,
    "tomador_razao_social": 40,
    "tomador_cnpj": 22,
    "tomador_cpf": 18,
    "data_emissao": 20,
    "tomador_email": 32,
    "tomador_logradouro": 32,
    "tomador_numero": 14,
    "tomador_complemento": 16,
    "tomador_bairro": 20,
    "tomador_cep": 14,
    "tomador_municipio_ibge": 22,
    "tomador_uf": 10,
    "deducoes": 16,
    "iss_retido": 14,
    "optante_simples": 22,
    "natureza_operacao": 22,
    "codigo_cnae": 14,
    "tomador_inscricao_municipal": 26,
}

_COLUNAS_RESULTADO = ["status", "numero_nfse", "codigo_verificacao", "link_consulta", "erro"]
_DESCRICOES_RESULTADO: dict[str, str] = {
    "status": "Status",
    "numero_nfse": "Número NFS-e",
    "codigo_verificacao": "Código Verificação",
    "link_consulta": "Link Consulta",
    "erro": "Mensagem de Erro",
}


def _preencher_aba(ws, colunas: list, exemplo: dict, larguras_extra: dict | None = None) -> None:
    """Preenche uma aba com cabeçalhos, linha de exemplo e formatação."""
    larguras = {**_LARGURAS, **(larguras_extra or {})}
    todas = colunas + _COLUNAS_RESULTADO

    for col_idx, col_name in enumerate(todas, start=1):
        col_letra = get_column_letter(col_idx)
        eh_resultado = col_name in _COLUNAS_RESULTADO
        eh_obrigatorio = "*" in DESCRICOES.get(col_name, "")

        cel = ws.cell(row=1, column=col_idx)
        if eh_resultado:
            descricao = _DESCRICOES_RESULTADO[col_name]
            cor_fundo = _COR_RESULTADO
        else:
            descricao = DESCRICOES.get(col_name, col_name)
            cor_fundo = _COR_OBRIGATORIO if eh_obrigatorio else _COR_OPCONAL

        cel.value = descricao
        cel.font = Font(bold=True, color="FFFFFF", size=9)
        cel.fill = PatternFill("solid", fgColor=cor_fundo)
        cel.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

        if col_name in exemplo and not eh_resultado:
            cel_ex = ws.cell(row=2, column=col_idx)
            cel_ex.value = exemplo[col_name]
            cel_ex.fill = PatternFill("solid", fgColor=_COR_EXEMPLO)
            cel_ex.alignment = Alignment(vertical="center")

        ws.column_dimensions[col_letra].width = larguras.get(col_name, 18)

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A2"


def criar_template_campinas(caminho: Path) -> None:
    """Gera o template Excel para emissão de NFS-e em Campinas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "NFS-e Campinas"

    _preencher_aba(ws, TODAS_COLUNAS, EXEMPLO)

    ws_inst = wb.create_sheet("Instruções")
    _preencher_instrucoes(ws_inst)

    wb.save(caminho)


def criar_template_sp(caminho: Path) -> None:
    """Gera o template Excel para emissão de NFS-e na Prefeitura de São Paulo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "NFS-e SP"

    _preencher_aba(ws, TODAS_COLUNAS_SP, EXEMPLO_SP, larguras_extra=_LARGURAS_SP_EXTRAS)

    ws_inst = wb.create_sheet("Instruções")
    _preencher_instrucoes_sp(ws_inst)

    wb.save(caminho)


def _preencher_instrucoes_sp(ws) -> None:
    instrucoes = [
        ("INSTRUÇÕES DE PREENCHIMENTO — NFS-e SÃO PAULO", True),
        ("", False),
        ("CAMPOS OBRIGATÓRIOS (marcados com *)", True),
        ("numero_rps        Número sequencial do RPS. Cada linha deve ter um número único.", False),
        ("competencia       Mês/ano de competência no formato MM/AAAA. Exemplo: 01/2026", False),
        ("discriminacao     Descrição completa do serviço prestado.", False),
        ("codigo_servico    Código de 5 dígitos do sistema ISS-SP. Ex: 07498 (consultoria TI).", False),
        ("valor_servico     Valor dos serviços em reais. Use ponto como separador decimal.", False),
        ("tomador_razao_social  Nome ou razão social do tomador.", False),
        ("tomador_cnpj      CNPJ do tomador (PJ). OU", False),
        ("tomador_cpf       CPF do tomador (PF). Preencha um dos dois.", False),
        ("aliquota_servicos Alíquota do ISS em porcentagem. Ex: 5.00 para 5%. Consulte a tabela SP.", False),
        ("", False),
        ("CAMPOS OPCIONAIS", True),
        ("data_emissao      Data de emissão no formato DD/MM/AAAA. Se vazio, usa a data de hoje.", False),
        ("iss_retido        ISS retido na fonte? S ou N. Default: N", False),
        ("optante_simples   Tomador é optante pelo Simples Nacional? S ou N. Default: N", False),
        ("deducoes          Valor das deduções em reais. Default: 0", False),
        ("tributacao_rps    T=Tributado (default), F=Fixo, J=Isento, A=Exterior, B=Imune, M=Fora SP", False),
        ("", False),
        ("RESULTADO (preenchido automaticamente após emissão)", True),
        ("status            EMITIDA ou ERRO", False),
        ("numero_nfse       Número da NFS-e emitida.", False),
        ("codigo_verificacao  Código de verificação para consulta.", False),
        ("erro              Mensagem de erro, se houver.", False),
        ("", False),
        ("DICAS", True),
        ("- Mantenha a linha 1 (cabeçalho) sem alterações.", False),
        ("- A linha 2 é apenas um exemplo; substitua pelos dados reais.", False),
        ("- Adicione uma linha por nota fiscal a emitir.", False),
        ("- Código IBGE de São Paulo: 3550308", False),
        ("- Os códigos de serviço SP (5 dígitos) são diferentes do LC 116/2003.", False),
        ("  Consulte o portal NFS-e SP para obter o código correto.", False),
    ]
    ws.column_dimensions["A"].width = 90
    for i, (texto, negrito) in enumerate(instrucoes, start=1):
        cel = ws.cell(row=i, column=1, value=texto)
        if negrito:
            cel.font = Font(bold=True)


def _preencher_instrucoes(ws) -> None:
    instrucoes = [
        ("INSTRUÇÕES DE PREENCHIMENTO — NFS-e CAMPINAS", True),
        ("", False),
        ("CAMPOS OBRIGATÓRIOS (marcados com *)", True),
        ("numero_rps        Número sequencial do RPS. Cada linha deve ter um número único.", False),
        ("competencia       Mês/ano de competência no formato MM/AAAA. Exemplo: 01/2026", False),
        ("discriminacao     Descrição completa do serviço prestado.", False),
        ("codigo_servico    Código LC 116/2003. Exemplos: 1.07 (informática), 17.01 (assessoria).", False),
        ("valor_servico     Valor dos serviços em reais. Use ponto como separador decimal.", False),
        ("tomador_razao_social  Nome ou razão social do tomador (cliente final).", False),
        ("tomador_cnpj      CNPJ do tomador (pessoa jurídica). OU", False),
        ("tomador_cpf       CPF do tomador (pessoa física). Preencha um dos dois.", False),
        ("", False),
        ("CAMPOS OPCIONAIS", True),
        ("data_emissao      Data de emissão no formato DD/MM/AAAA. Se vazio, usa a data de hoje.", False),
        ("iss_retido        ISS retido na fonte? S ou N. Default: N", False),
        ("optante_simples   Tomador é optante pelo Simples Nacional? S ou N. Default: N", False),
        ("deducoes          Valor das deduções em reais. Default: 0", False),
        ("natureza_operacao 1=Tributação no município (default), 2=Fora do município", False),
        ("", False),
        ("RESULTADO (preenchido automaticamente após emissão)", True),
        ("status            EMITIDA ou ERRO", False),
        ("numero_nfse       Número da NFS-e emitida.", False),
        ("codigo_verificacao  Código de verificação para consulta.", False),
        ("link_consulta     URL para visualização/download da nota.", False),
        ("erro              Mensagem de erro, se houver.", False),
        ("", False),
        ("DICAS", True),
        ("- Mantenha a linha 1 (cabeçalho) sem alterações.", False),
        ("- A linha 2 é apenas um exemplo; substitua pelos dados reais.", False),
        ("- Adicione uma linha por nota fiscal a emitir.", False),
        ("- Código IBGE de Campinas: 3509502", False),
        ("- Códigos de serviço: consulte o ISS Campinas para o código municipal correto.", False),
    ]
    ws.column_dimensions["A"].width = 90
    for i, (texto, negrito) in enumerate(instrucoes, start=1):
        cel = ws.cell(row=i, column=1, value=texto)
        if negrito:
            cel.font = Font(bold=True)
